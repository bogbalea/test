from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd
import os
from openpyxl import load_workbook

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def load_workbook_once(output_file):
    return load_workbook(output_file)


@app.route('/', methods=['GET', 'POST'])
def upload_files():
    if request.method == 'POST':
        # Save the Historic Cost File
        historic_file = request.files['historic_cost']
        historic_path = os.path.join(UPLOAD_FOLDER, historic_file.filename)
        historic_file.save(historic_path)

        # Save up to four Price List files
        price_list_paths = []
        for i in range(1, 5):
            price_list_file = request.files.get(f'price_list_{i}')
            if price_list_file:
                path = os.path.join(UPLOAD_FOLDER, price_list_file.filename)
                price_list_file.save(path)
                price_list_paths.append(path)

        # Redirect to processing route
        return redirect(url_for('process_files', historic_file=historic_file.filename,
                                price_list_files=','.join([os.path.basename(p) for p in price_list_paths])))

    return render_template('upload_new.html')


# Helper function to calculate the cost and return the price used for the calculation
def calculate_cost_and_price(row, price_df, rate_type):
    # Construct the lane based on each Rate Geography type
    if rate_type == 'CITY-CITY':
        lane = f"{row['Source City']}-{row['Destination City']}"
    elif rate_type == 'POSTAL-POSTAL':
        lane = f"{row['Source Postal']}-{row['Destination Postal']}"
    elif rate_type == 'REGION-REGION':
        lane = f"{row['Source Region']}-{row['Destination Region']}"
    elif rate_type == 'LOCATION-LOCATION':
        lane = f"{row['Source Location']}-{row['Destination Location']}"
    elif rate_type == 'CITY-POSTAL':
        lane = f"{row['Source City']}-{row['Destination Postal']}"
    elif rate_type == 'CITY-REGION':
        lane = f"{row['Source City']}-{row['Destination Region']}"
    elif rate_type == 'CITY-LOCATION':
        lane = f"{row['Source City']}-{row['Destination Location']}"
    elif rate_type == 'POSTAL-CITY':
        lane = f"{row['Source Postal']}-{row['Destination City']}"
    elif rate_type == 'POSTAL-REGION':
        lane = f"{row['Source Postal']}-{row['Destination Region']}"
    elif rate_type == 'POSTAL-LOCATION':
        lane = f"{row['Source Postal']}-{row['Destination Location']}"
    elif rate_type == 'REGION-CITY':
        lane = f"{row['Source Region']}-{row['Destination City']}"
    elif rate_type == 'REGION-POSTAL':
        lane = f"{row['Source Region']}-{row['Destination Postal']}"
    elif rate_type == 'REGION-LOCATION':
        lane = f"{row['Source Region']}-{row['Destination Location']}"
    elif rate_type == 'LOCATION-CITY':
        lane = f"{row['Source Location']}-{row['Destination City']}"
    elif rate_type == 'LOCATION-POSTAL':
        lane = f"{row['Source Location']}-{row['Destination Postal']}"
    elif rate_type == 'LOCATION-REGION':
        lane = f"{row['Source Location']}-{row['Destination Region']}"
    else:
        lane = "UNKNOWN"  # Default for unsupported or unexpected Rate Geographies

    # Filter the price based on Lane and Adjusted Cost Type
    price_row = price_df[(price_df['Lane'] == lane) & (price_df['Adjusted Cost Type'] == row['Adjusted Cost Type'])]

    # Return the calculated cost and the price used
    if not price_row.empty:
        price = price_row.iloc[0]['Price']
        if row['Cost Type'] == 'KM':
            return row['Traveled Distance'] * price, price
        elif row['Cost Type'] == 'KG':
            return row['Weight'] * price, price
        else:
            return price, price
    return None, None



# Function to create a summary of costs for each carrier using dynamically generated carrier names
def create_carrier_summary(output_file, df):
    # Load the workbook and access the main sheet (where recalculated costs are saved)
    wb = load_workbook(output_file)
    main_sheet = wb.active  # Assuming the main sheet is the active one

    # Create a new sheet for the summary
    if "Scenario1-Best_Carrier" in wb.sheetnames:
        del wb["Scenario1-Best_Carrier"]  # Remove if it exists (optional)
    ws = wb.create_sheet("Scenario1-Best_Carrier")  # Renamed to Scenario1-Best_Carrier

    # Set up headers for Historic Cost, Carrier Name, Recalculated Cost, Difference, Rank, and % Difference
    ws['B1'] = 'Total Historic Cost'
    ws['C1'] = 'Carrier Name'
    ws['D1'] = 'Total Recalculated Cost'
    ws['E1'] = 'Total Difference'
    ws['F1'] = 'Rank'
    ws['G1'] = '% Difference'

    # Sum the Historic Cost (assuming 'Total Cost' exists in the DataFrame as 'Total Cost' column)
    total_historic_cost = df['Total Cost'].sum()
    ws['B2'] = total_historic_cost

    # Get dynamically generated carrier names from the DataFrame (assuming they have '_Recalculated_Cost' suffix)
    carrier_columns = [col for col in df.columns if '_Recalculated_Cost' in col]

    # Prepare a list to store carrier data (to rank them later)
    carrier_data = []

    # Loop through the carrier columns to calculate the total recalculated cost and difference
    for idx, col_name in enumerate(carrier_columns, start=2):
        carrier_name = col_name.replace('_Recalculated_Cost', '')  # Extract carrier name from column name
        total_recalculated_cost = df[col_name].sum()  # Sum the recalculated costs for each carrier

        # Calculate the difference between the total historic cost and recalculated cost
        total_difference = total_recalculated_cost - total_historic_cost

        # Calculate % Difference (Total Difference / Total Historic Cost)
        percent_difference = (total_difference / total_historic_cost) * 100

        # Store the data for later ranking
        carrier_data.append({
            'carrier_name': carrier_name,
            'total_recalculated_cost': total_recalculated_cost,
            'total_difference': total_difference,
            'percent_difference': percent_difference
        })

    # Rank the carriers based on the total recalculated cost (ascending order)
    carrier_data = sorted(carrier_data, key=lambda x: x['total_recalculated_cost'])

    # Populate the summary sheet with ranked data
    for idx, carrier in enumerate(carrier_data, start=2):
        ws[f'C{idx}'] = carrier['carrier_name']  # Carrier Name
        ws[f'D{idx}'] = carrier['total_recalculated_cost']  # Total Recalculated Cost
        ws[f'E{idx}'] = carrier['total_difference']  # Total Difference
        ws[f'F{idx}'] = idx - 1  # Rank
        ws[f'G{idx}'] = carrier['percent_difference']  # % Difference

    # Save the workbook with the new summary sheet
    wb.save(output_file)

# Function to create a summary for "Scenario2-BC_Simulation" using Minimum_Simulation_Cost
def create_carrier_simulation_summary(output_file, df):
    # Load the workbook and access the main sheet (where recalculated costs are saved)
    wb = load_workbook(output_file)
    main_sheet = wb.active  # Assuming the main sheet is the active one

    # Create a new sheet for the summary
    if "Scenario2-BC_Simulation" in wb.sheetnames:
        del wb["Scenario2-BC_Simulation"]  # Remove if it exists (optional)
    ws = wb.create_sheet("Scenario2-BC_Simulation")  # Renamed to Scenario2-BC_Simulation

    # Set up headers for Historic Cost, Carrier Name, Minimum Simulation Cost, Difference, Rank, and % Difference
    ws['B1'] = 'Total Historic Cost'
    ws['C1'] = 'Carrier Name'
    ws['D1'] = 'Minimum Simulation Cost'
    ws['E1'] = 'Total Difference'
    ws['F1'] = 'Rank'
    ws['G1'] = '% Difference'

    # Sum the Historic Cost (assuming 'Total Cost' exists in the DataFrame as 'Total Cost' column)
    total_historic_cost = df['Total Cost'].sum()
    ws['B2'] = total_historic_cost

    # Get dynamically generated carrier names from the DataFrame (assuming they have '_Minimum_Simulation_Cost' suffix)
    carrier_columns = [col for col in df.columns if '_Minimum_Simulation_Cost' in col]

    # Prepare a list to store carrier data (to rank them later)
    carrier_data = []

    # Loop through the carrier columns to calculate the total Minimum Simulation Cost and difference
    for idx, col_name in enumerate(carrier_columns, start=2):
        carrier_name = col_name.replace('_Minimum_Simulation_Cost', '')  # Extract carrier name from column name
        total_simulation_cost = df[col_name].sum()  # Sum the Minimum Simulation Costs for each carrier

        # Calculate the difference between the total historic cost and minimum simulation cost
        total_difference = total_simulation_cost - total_historic_cost

        # Calculate % Difference (Total Difference / Total Historic Cost)
        percent_difference = (total_difference / total_historic_cost) * 100

        # Store the data for later ranking
        carrier_data.append({
            'carrier_name': carrier_name,
            'total_simulation_cost': total_simulation_cost,
            'total_difference': total_difference,
            'percent_difference': percent_difference
        })

    # Rank the carriers based on the total simulation cost (ascending order)
    carrier_data = sorted(carrier_data, key=lambda x: x['total_simulation_cost'])

    # Populate the summary sheet with ranked data
    for idx, carrier in enumerate(carrier_data, start=2):
        ws[f'C{idx}'] = carrier['carrier_name']  # Carrier Name
        ws[f'D{idx}'] = carrier['total_simulation_cost']  # Minimum Simulation Cost
        ws[f'E{idx}'] = carrier['total_difference']  # Total Difference
        ws[f'F{idx}'] = idx - 1  # Rank
        ws[f'G{idx}'] = carrier['percent_difference']  # % Difference

    # Save the workbook with the new summary sheet
    wb.save(output_file)


# Function to create a summary for "Scenario3-Best_Mode" with a breakdown by Mode using Total Recalculated Cost
def create_mode_summary(output_file, df):
    # Load the workbook and access the main sheet (where recalculated costs are saved)
    wb = load_workbook(output_file)
    main_sheet = wb.active  # Assuming the main sheet is the active one

    # Create a new sheet for the summary
    if "Scenario3-Best_Mode" in wb.sheetnames:
        del wb["Scenario3-Best_Mode"]  # Remove if it exists (optional)
    ws = wb.create_sheet("Scenario3-Best_Mode")  # New sheet for Best Mode breakdown

    # Set up headers for Mode, Historic Cost, Carrier Name, Total Recalculated Cost, Difference, Rank, and % Difference
    ws['B1'] = 'Mode'
    ws['C1'] = 'Total Historic Cost'
    ws['D1'] = 'Carrier Name'
    ws['E1'] = 'Total Recalculated Cost'
    ws['F1'] = 'Total Difference'
    ws['G1'] = 'Rank'
    ws['H1'] = '% Difference'

    # Group by Mode and perform calculations for each mode
    unique_modes = df['Mode'].unique()

    row_idx = 2  # Start populating from row 2
    for mode in unique_modes:
        mode_df = df[df['Mode'] == mode]  # Filter the DataFrame by Mode

        # Sum the Historic Cost for the current mode
        total_historic_cost = mode_df['Total Cost'].sum()

        # Get dynamically generated carrier names from the DataFrame (assuming they have '_Recalculated_Cost' suffix)
        carrier_columns = [col for col in mode_df.columns if '_Recalculated_Cost' in col]

        # Prepare a list to store carrier data (to rank them later)
        carrier_data = []

        # Loop through the carrier columns to calculate the total Recalculated Cost and difference
        for col_name in carrier_columns:
            carrier_name = col_name.replace('_Recalculated_Cost', '')  # Extract carrier name from column name
            total_recalculated_cost = mode_df[col_name].sum()  # Sum the Total Recalculated Costs for each carrier

            # Calculate the difference between the total historic cost and recalculated cost
            total_difference = total_recalculated_cost - total_historic_cost

            # Calculate % Difference (Total Difference / Total Historic Cost)
            percent_difference = (total_difference / total_historic_cost) * 100

            # Store the data for later ranking
            carrier_data.append({
                'carrier_name': carrier_name,
                'total_recalculated_cost': total_recalculated_cost,
                'total_difference': total_difference,
                'percent_difference': percent_difference
            })

        # Rank the carriers based on the total recalculated cost (ascending order)
        carrier_data = sorted(carrier_data, key=lambda x: x['total_recalculated_cost'])

        # Populate the summary sheet with ranked data for the current mode
        mode_filled = False  # Used to control when to fill the mode label
        cost_filled = False  # Used to control when to fill the historic cost
        for idx, carrier in enumerate(carrier_data):
            # Only write the mode for the first row of each group
            if not mode_filled:
                ws[f'B{row_idx}'] = mode  # Mode
                mode_filled = True  # Mode label filled once per group

            # Only write the total historic cost for the first row of each group
            if not cost_filled:
                ws[f'C{row_idx}'] = total_historic_cost  # Total Historic Cost
                cost_filled = True  # Total Historic Cost filled once per group

            ws[f'D{row_idx}'] = carrier['carrier_name']  # Carrier Name
            ws[f'E{row_idx}'] = carrier['total_recalculated_cost']  # Total Recalculated Cost
            ws[f'F{row_idx}'] = carrier['total_difference']  # Total Difference
            ws[f'G{row_idx}'] = idx + 1  # Rank
            ws[f'H{row_idx}'] = carrier['percent_difference']  # % Difference
            row_idx += 1

    # Save the workbook with the new summary sheet
    wb.save(output_file)

# Function to create a summary for "Scenario4-BM_Simulation" with a breakdown by Mode using Minimum Simulation Cost
def create_bm_simulation_summary(output_file, df):
    # Load the workbook and access the main sheet (where recalculated costs are saved)
    wb = load_workbook(output_file)
    main_sheet = wb.active  # Assuming the main sheet is the active one

    # Create a new sheet for the summary
    if "Scenario4-BM_Simulation" in wb.sheetnames:
        del wb["Scenario4-BM_Simulation"]  # Remove if it exists (optional)
    ws = wb.create_sheet("Scenario4-BM_Simulation")  # New sheet for BM Simulation breakdown

    # Set up headers for Mode, Historic Cost, Carrier Name, Minimum Simulation Cost, Difference, Rank, and % Difference
    ws['B1'] = 'Mode'
    ws['C1'] = 'Total Historic Cost'
    ws['D1'] = 'Carrier Name'
    ws['E1'] = 'Minimum Simulation Cost'
    ws['F1'] = 'Total Difference'
    ws['G1'] = 'Rank'
    ws['H1'] = '% Difference'

    # Group by Mode and perform calculations for each mode
    unique_modes = df['Mode'].unique()

    row_idx = 2  # Start populating from row 2
    for mode in unique_modes:
        mode_df = df[df['Mode'] == mode]  # Filter the DataFrame by Mode

        # Sum the Historic Cost for the current mode
        total_historic_cost = mode_df['Total Cost'].sum()

        # Get dynamically generated carrier names from the DataFrame (assuming they have '_Minimum_Simulation_Cost' suffix)
        carrier_columns = [col for col in mode_df.columns if '_Minimum_Simulation_Cost' in col]

        # Prepare a list to store carrier data (to rank them later)
        carrier_data = []

        # Loop through the carrier columns to calculate the total Minimum Simulation Cost and difference
        for col_name in carrier_columns:
            carrier_name = col_name.replace('_Minimum_Simulation_Cost', '')  # Extract carrier name from column name
            total_simulation_cost = mode_df[col_name].sum()  # Sum the Minimum Simulation Costs for each carrier

            # Calculate the difference between the total historic cost and minimum simulation cost
            total_difference = total_simulation_cost - total_historic_cost

            # Calculate % Difference (Total Difference / Total Historic Cost)
            percent_difference = (total_difference / total_historic_cost) * 100

            # Store the data for later ranking
            carrier_data.append({
                'carrier_name': carrier_name,
                'total_simulation_cost': total_simulation_cost,
                'total_difference': total_difference,
                'percent_difference': percent_difference
            })

        # Rank the carriers based on the total simulation cost (ascending order)
        carrier_data = sorted(carrier_data, key=lambda x: x['total_simulation_cost'])

        # Populate the summary sheet with ranked data for the current mode
        mode_filled = False  # Used to control when to fill the mode label
        cost_filled = False  # Used to control when to fill the historic cost
        for idx, carrier in enumerate(carrier_data):
            # Only write the mode for the first row of each group
            if not mode_filled:
                ws[f'B{row_idx}'] = mode  # Mode
                mode_filled = True  # Mode label filled once per group

            # Only write the total historic cost for the first row of each group
            if not cost_filled:
                ws[f'C{row_idx}'] = total_historic_cost  # Total Historic Cost
                cost_filled = True  # Total Historic Cost filled once per group

            ws[f'D{row_idx}'] = carrier['carrier_name']  # Carrier Name
            ws[f'E{row_idx}'] = carrier['total_simulation_cost']  # Minimum Simulation Cost
            ws[f'F{row_idx}'] = carrier['total_difference']  # Total Difference
            ws[f'G{row_idx}'] = idx + 1  # Rank
            ws[f'H{row_idx}'] = carrier['percent_difference']  # % Difference
            row_idx += 1

    # Save the workbook with the new summary sheet
    wb.save(output_file)


# Function to create a summary for "Scenario5-Best_Movement" with a breakdown by Movement using Total Recalculated Cost
def create_best_movement_summary(output_file, df):
    wb = load_workbook(output_file)
    main_sheet = wb.active

    if "Scenario5-Best_Movement" in wb.sheetnames:
        del wb["Scenario5-Best_Movement"]
    ws = wb.create_sheet("Scenario5-Best_Movement")

    ws['B1'] = 'Movement'
    ws['C1'] = 'Total Historic Cost'
    ws['D1'] = 'Carrier Name'
    ws['E1'] = 'Total Recalculated Cost'
    ws['F1'] = 'Total Difference'
    ws['G1'] = 'Rank'
    ws['H1'] = '% Difference'

    unique_movements = df['Movement'].unique()
    row_idx = 2

    for movement in unique_movements:
        movement_df = df[df['Movement'] == movement]
        total_historic_cost = movement_df['Total Cost'].sum()
        carrier_columns = [col for col in movement_df.columns if '_Recalculated_Cost' in col]

        carrier_data = []
        for col_name in carrier_columns:
            carrier_name = col_name.replace('_Recalculated_Cost', '')
            total_recalculated_cost = movement_df[col_name].sum()
            total_difference = total_recalculated_cost - total_historic_cost
            percent_difference = (total_difference / total_historic_cost) * 100

            carrier_data.append({
                'carrier_name': carrier_name,
                'total_recalculated_cost': total_recalculated_cost,
                'total_difference': total_difference,
                'percent_difference': percent_difference
            })

        carrier_data = sorted(carrier_data, key=lambda x: x['total_recalculated_cost'])

        movement_filled = False
        cost_filled = False
        for idx, carrier in enumerate(carrier_data):
            if not movement_filled:
                ws[f'B{row_idx}'] = movement
                movement_filled = True
            if not cost_filled:
                ws[f'C{row_idx}'] = total_historic_cost
                cost_filled = True

            ws[f'D{row_idx}'] = carrier['carrier_name']
            ws[f'E{row_idx}'] = carrier['total_recalculated_cost']
            ws[f'F{row_idx}'] = carrier['total_difference']
            ws[f'G{row_idx}'] = idx + 1
            ws[f'H{row_idx}'] = carrier['percent_difference']
            row_idx += 1

    wb.save(output_file)


# Function to create a summary for "Scenario6-BMV_Simulation" with a breakdown by Movement using Minimum Simulation Cost
def create_bmv_simulation_summary(output_file, df):
    wb = load_workbook(output_file)
    main_sheet = wb.active

    if "Scenario6-BMV_Simulation" in wb.sheetnames:
        del wb["Scenario6-BMV_Simulation"]
    ws = wb.create_sheet("Scenario6-BMV_Simulation")

    ws['B1'] = 'Movement'
    ws['C1'] = 'Total Historic Cost'
    ws['D1'] = 'Carrier Name'
    ws['E1'] = 'Minimum Simulation Cost'
    ws['F1'] = 'Total Difference'
    ws['G1'] = 'Rank'
    ws['H1'] = '% Difference'

    unique_movements = df['Movement'].unique()
    row_idx = 2

    for movement in unique_movements:
        movement_df = df[df['Movement'] == movement]
        total_historic_cost = movement_df['Total Cost'].sum()
        carrier_columns = [col for col in movement_df.columns if '_Minimum_Simulation_Cost' in col]

        carrier_data = []
        for col_name in carrier_columns:
            carrier_name = col_name.replace('_Minimum_Simulation_Cost', '')
            total_simulation_cost = movement_df[col_name].sum()
            total_difference = total_simulation_cost - total_historic_cost
            percent_difference = (total_difference / total_historic_cost) * 100

            carrier_data.append({
                'carrier_name': carrier_name,
                'total_simulation_cost': total_simulation_cost,
                'total_difference': total_difference,
                'percent_difference': percent_difference
            })

        carrier_data = sorted(carrier_data, key=lambda x: x['total_simulation_cost'])

        movement_filled = False
        cost_filled = False
        for idx, carrier in enumerate(carrier_data):
            if not movement_filled:
                ws[f'B{row_idx}'] = movement
                movement_filled = True
            if not cost_filled:
                ws[f'C{row_idx}'] = total_historic_cost
                cost_filled = True

            ws[f'D{row_idx}'] = carrier['carrier_name']
            ws[f'E{row_idx}'] = carrier['total_simulation_cost']
            ws[f'F{row_idx}'] = carrier['total_difference']
            ws[f'G{row_idx}'] = idx + 1
            ws[f'H{row_idx}'] = carrier['percent_difference']
            row_idx += 1

    wb.save(output_file)



# Processing Route - Recalculation and Simulation
@app.route('/process/<historic_file>/<price_list_files>')
def process_files(historic_file, price_list_files):
    historic_path = os.path.join(UPLOAD_FOLDER, historic_file)
    price_list_files = price_list_files.split(',')

    # Load the Historic Cost file
    historic_df = pd.read_excel(historic_path)

    # Define custom column names based on the observed structure
    custom_column_names = [
        'Shipment ID', 'Shipment Date', 'Carrier', 'Mode', 'Movement',
        'Source Location', 'Source City', 'Source Postal', 'Source Region',
        'Destination Location', 'Destination City', 'Destination Postal',
        'Destination Region', 'Weight', 'Volume', 'Loading Meters',
        'Truck Type', 'Traveled Distance', 'Total Cost', 'Rate Geography',
        'Cost Type', 'Remarks', 'User Own Reference'
    ]
    historic_df.columns = custom_column_names

    # Create Adjusted Cost Type based on Truck Type for EQUIPMENT
    historic_df['Adjusted Cost Type'] = historic_df.apply(
        lambda row: row['Truck Type'] if row['Cost Type'] == 'EQUIPMENT' else row['Cost Type'],
        axis=1
    )

    # Load the price list files into separate DataFrames
    price_df_list = [pd.read_excel(os.path.join(UPLOAD_FOLDER, price_file)) for price_file in price_list_files]

    all_results = []
    for i, price_df in enumerate(price_df_list):
        carrier_name = f"carrier{i + 1}"
        price_df['Lane'] = price_df['SOURCE'] + '-' + price_df['DESTINATION']
        price_df['Adjusted Cost Type'] = price_df.apply(
            lambda row: row['Truck Type'] if row['Cost Type'] == 'EQUIPMENT' else row['Cost Type'],
            axis=1
        )

        # Part 1: Calculate Recalculated Cost and Price Using the Original Rate Geography
        recalculated_cost, recalculated_price = zip(*historic_df.apply(
            lambda row: calculate_cost_and_price(row, price_df, row['Rate Geography']),
            axis=1
        ))
        historic_df[f'{carrier_name}_Recalculated_Cost'] = recalculated_cost
        historic_df[f'{carrier_name}_Recalculated_Price'] = recalculated_price

        # Calculate the difference between the historic cost and the recalculated cost
        historic_df[f'{carrier_name}_Difference'] = historic_df[f'{carrier_name}_Recalculated_Cost'] - historic_df['Total Cost']

        # Calculate the percentage difference between the recalculated cost and the historic cost
        historic_df[f'{carrier_name}_Recalculated_Percentage'] = (historic_df[f'{carrier_name}_Difference'] / historic_df['Total Cost']) * 100

        # Part 2: Simulation - Calculate costs and prices for the remaining 15 geography types
        simulation_results = {}
        for rate_type in ['CITY-CITY', 'POSTAL-POSTAL', 'REGION-REGION', 'LOCATION-LOCATION',
                          'CITY-POSTAL', 'CITY-REGION', 'CITY-LOCATION', 'POSTAL-CITY', 'POSTAL-REGION',
                          'POSTAL-LOCATION', 'REGION-CITY', 'REGION-POSTAL', 'REGION-LOCATION',
                          'LOCATION-CITY', 'LOCATION-POSTAL']:
            sim_col_name = f'{carrier_name}_{rate_type}_Simulation'
            price_col_name = f'{carrier_name}_{rate_type}_Price'
            simulation_cost, simulation_price = zip(*historic_df.apply(
                lambda row: calculate_cost_and_price(row, price_df, rate_type),
                axis=1
            ))
            simulation_results[sim_col_name] = simulation_cost
            simulation_results[price_col_name] = simulation_price

        # Concat the results for better performance
        simulation_df = pd.DataFrame(simulation_results)
        historic_df = pd.concat([historic_df, simulation_df], axis=1)

        # Create Minimum Simulation Cost and Price for the current carrier
        simulation_columns = [col for col in simulation_results if col.endswith('_Simulation')]
        historic_df[f'{carrier_name}_Minimum_Simulation_Cost'] = historic_df[simulation_columns].min(axis=1)

        # Calculate the difference between the historic cost and the minimum simulation cost
        historic_df[f'{carrier_name}_Simulation_Difference'] = historic_df[f'{carrier_name}_Minimum_Simulation_Cost'] - historic_df['Total Cost']

        # Calculate the percentage difference between the simulation cost and the historic cost
        historic_df[f'{carrier_name}_Simulation_Percentage'] = (historic_df[f'{carrier_name}_Simulation_Difference'] / historic_df['Total Cost']) * 100

        # Find the corresponding minimum price
        historic_df[f'{carrier_name}_Minimum_Simulation_Price'] = historic_df.apply(
            lambda row: [historic_df[col.replace('_Simulation', '_Price')][row.name] for col in simulation_columns if
                         row[col] == row[f'{carrier_name}_Minimum_Simulation_Cost']][0]
            if pd.notnull(row[f'{carrier_name}_Minimum_Simulation_Cost']) and
               len([col.replace('_Simulation', '_Price') for col in simulation_columns if
                    row[col] == row[f'{carrier_name}_Minimum_Simulation_Cost']]) > 0
            else None,
            axis=1
        )

        # Create Minimum Rate Geography for the current carrier
        historic_df[f'{carrier_name}_Minimum_Rate_Geography'] = historic_df.apply(
            lambda row: [col.split('_')[1] for col in simulation_columns if
                         pd.notnull(row[col]) and row[col] == row[f'{carrier_name}_Minimum_Simulation_Cost']][0]
            if pd.notnull(row[f'{carrier_name}_Minimum_Simulation_Cost']) and
               len([col.split('_')[1] for col in simulation_columns if
                    pd.notnull(row[col]) and row[col] == row[f'{carrier_name}_Minimum_Simulation_Cost']]) > 0
            else None,
            axis=1
        )

    # Drop the simulation columns just before saving
    columns_to_drop = []
    for i in range(1, len(price_list_files) + 1):
        carrier_name = f"carrier{i}"
        for rate_type in ['CITY-CITY', 'POSTAL-POSTAL', 'REGION-REGION', 'LOCATION-LOCATION',
                          'CITY-POSTAL', 'CITY-REGION', 'CITY-LOCATION', 'POSTAL-CITY', 'POSTAL-REGION',
                          'POSTAL-LOCATION', 'REGION-CITY', 'REGION-POSTAL', 'REGION-LOCATION',
                          'LOCATION-CITY', 'LOCATION-POSTAL']:
            columns_to_drop.append(f'{carrier_name}_{rate_type}_Simulation')
            columns_to_drop.append(f'{carrier_name}_{rate_type}_Price')

    historic_df = historic_df.drop(columns=columns_to_drop, errors='ignore')

    # Save to Excel with the new columns per carrier
    output_file = os.path.join(UPLOAD_FOLDER, 'Final_Historic_Cost_per_Carrier_with_Prices.xlsx')
    historic_df.to_excel(output_file, index=False)

    # Create a summary of total costs per carrier (using dynamically generated carrier names)
    create_carrier_summary(output_file, historic_df)

    # Create the "Scenario2-BC_Simulation" summary using Minimum Simulation Cost
    create_carrier_simulation_summary(output_file, historic_df)

    # Create the "Scenario3-Best_Mode" summary with a breakdown by Mode
    create_mode_summary(output_file, historic_df)

    # Create the "Scenario4-BM_Simulation" summary with a breakdown by Mode using Minimum Simulation Cost
    create_bm_simulation_summary(output_file, historic_df)

    # Create the

    create_best_movement_summary(output_file, historic_df)
    create_bmv_simulation_summary(output_file, historic_df)

    return render_template('results_new.html', output_file='Final_Historic_Cost_per_Carrier_with_Prices.xlsx')

# Download Route for the updated file
@app.route('/download/<output_file>')
def download_file(output_file):
    path = os.path.join(UPLOAD_FOLDER, output_file)
    return send_file(path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
